"""Generate a realistic customers.xlsx for the v27 E2E upload demo."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

wb = Workbook()
ws = wb.active
ws.title = "Customers"

headers = ["name", "city", "plan", "ltv", "active"]
ws.append(headers)
rows = [
    ["Ada Obi", "Lagos", "pro", 1240.5, True],
    ["Grace Chen", "Berlin", "enterprise", 9800.0, True],
    ["Linus Park", "Seoul", "free", 0.0, False],
    ["Amara Diallo", "Dakar", "pro", 2110.75, True],
    ["Nadia Rossi", "Milan", "starter", 310.0, True],
    ["Ken Watanabe", "Osaka", "enterprise", 12450.0, True],
    ["Femi Adebayo", "Lagos", "starter", 285.25, False],
    ["Zoe Martin", "Paris", "pro", 1985.0, True],
]
for r in rows:
    ws.append(r)

head_fill = PatternFill("solid", fgColor="0F766E")
for cell in ws[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = head_fill
for col, width in zip("ABCDE", [18, 12, 12, 10, 9]):
    ws.column_dimensions[col].width = width

wb.save("/home/z/my-project/scripts/e2e-customers.xlsx")
print("wrote scripts/e2e-customers.xlsx with", len(rows), "rows")
